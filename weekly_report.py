#!/usr/bin/env python3
# ==========================================================
# WEEKLY TENDER DASHBOARD REPORT
# Generates comprehensive weekly summary with analytics
# ==========================================================

import os
import sys
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import yaml

# Load config
config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.yaml")
with open(config_path, "r") as f:
    CONFIG = yaml.safe_load(f)

EXCEL_PATH = CONFIG["paths"]["tender_log_excel"]
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")
REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")

# Email settings
EMAIL_ENABLED = os.environ.get("TENDERSCAN_EMAIL_ENABLED", "false").lower() == "true"
SMTP_SERVER = os.environ.get("TENDERSCAN_SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("TENDERSCAN_SMTP_PORT", "587"))
SMTP_USER = os.environ.get("TENDERSCAN_SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("TENDERSCAN_SMTP_PASSWORD", "")
EMAIL_TO = os.environ.get("TENDERSCAN_EMAIL_TO", "").split(",")
EMAIL_FROM = os.environ.get("TENDERSCAN_EMAIL_FROM", SMTP_USER)


def get_weekly_stats():
    """Extract stats from Excel for the past week"""
    
    if not os.path.exists(EXCEL_PATH):
        return None
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    week_ago = datetime.now() - timedelta(days=7)
    
    stats = {
        "total": 0,
        "this_week": 0,
        "by_type": {"TES": 0, "Phakathi": 0, "Both": 0, "Unknown": 0},
        "by_priority": {"HIGH": 0, "MEDIUM": 0, "LOW": 0},
        "by_status": {},
        "closing_soon": [],
        "high_priority": [],
        "top_industries": {}
    }
    
    for row in range(2, ws.max_row + 1):
        stats["total"] += 1
        
        tender_name = ws.cell(row=row, column=1).value or ""
        client = ws.cell(row=row, column=2).value or ""
        t_type = ws.cell(row=row, column=3).value or "Unknown"
        industry = ws.cell(row=row, column=4).value or "Unknown"
        composite = ws.cell(row=row, column=6).value or 5
        priority = ws.cell(row=row, column=7).value or "MEDIUM"
        closing = ws.cell(row=row, column=13).value or ""
        status = ws.cell(row=row, column=14).value or "Open"
        date_added = ws.cell(row=row, column=18).value or ""
        ref = ws.cell(row=row, column=17).value or ""
        
        if t_type in stats["by_type"]:
            stats["by_type"][t_type] += 1
        
        if priority in stats["by_priority"]:
            stats["by_priority"][priority] += 1
        
        stats["by_status"][status] = stats["by_status"].get(status, 0) + 1
        
        ind_key = str(industry).split("(")[0].strip()[:20]
        stats["top_industries"][ind_key] = stats["top_industries"].get(ind_key, 0) + 1
        
        try:
            added = datetime.strptime(str(date_added), "%Y-%m-%d")
            if added >= week_ago:
                stats["this_week"] += 1
        except:
            pass
        
        try:
            close_date = datetime.strptime(str(closing), "%Y-%m-%d")
            days_left = (close_date - datetime.now()).days
            if 0 <= days_left <= 7 and status == "Open":
                stats["closing_soon"].append({
                    "ref": ref,
                    "title": tender_name[:50],
                    "client": client,
                    "days_left": days_left,
                    "priority": priority
                })
        except:
            pass
        
        if priority == "HIGH" and status == "Open":
            stats["high_priority"].append({
                "ref": ref,
                "title": tender_name[:50],
                "client": client,
                "type": t_type,
                "score": composite
            })
    
    stats["closing_soon"] = sorted(stats["closing_soon"], key=lambda x: x["days_left"])[:10]
    stats["high_priority"] = sorted(stats["high_priority"], key=lambda x: x.get("score", 0), reverse=True)[:10]
    stats["top_industries"] = dict(sorted(stats["top_industries"].items(), key=lambda x: x[1], reverse=True)[:5])
    
    return stats


def generate_weekly_html(stats: dict) -> str:
    """Generate weekly dashboard HTML"""
    
    if not stats:
        return "<html><body><h1>No data available</h1></body></html>"
    
    now = datetime.now()
    week_ago = now - timedelta(days=7)
    
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 900px; margin: 0 auto; }}
            h1 {{ color: #1565C0; border-bottom: 3px solid #1565C0; padding-bottom: 10px; }}
            h2 {{ color: #2E7D32; margin-top: 30px; }}
            .dashboard {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }}
            .card {{ background: #f5f5f5; padding: 20px; border-radius: 8px; text-align: center; }}
            .card.highlight {{ background: #E3F2FD; border: 2px solid #1565C0; }}
            .card h3 {{ margin: 0; font-size: 2em; color: #1565C0; }}
            .card p {{ margin: 5px 0 0 0; color: #666; }}
            .priority-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; }}
            .priority-card {{ padding: 15px; border-radius: 8px; text-align: center; }}
            .priority-card.high {{ background: #FFEBEE; color: #C62828; }}
            .priority-card.medium {{ background: #FFF8E1; color: #F57F17; }}
            .priority-card.low {{ background: #E8F5E9; color: #2E7D32; }}
            .priority-card h3 {{ font-size: 2.5em; margin: 0; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
            th {{ background: #1565C0; color: white; padding: 12px; text-align: left; }}
            td {{ border: 1px solid #ddd; padding: 10px; }}
            tr:nth-child(even) {{ background: #f9f9f9; }}
            .urgent {{ background: #FFEBEE; }}
            .badge {{ display: inline-block; padding: 3px 8px; border-radius: 4px; font-size: 0.8em; font-weight: bold; }}
            .badge-high {{ background: #F44336; color: white; }}
            .badge-medium {{ background: #FFC107; color: black; }}
            .badge-tes {{ background: #2196F3; color: white; }}
            .badge-phakathi {{ background: #9C27B0; color: white; }}
            .chart-container {{ margin: 20px 0; }}
            .bar {{ height: 25px; background: #1565C0; margin: 5px 0; border-radius: 3px; color: white; padding: 3px 10px; }}
        </style>
    </head>
    <body>
        <h1>📊 Weekly Tender Dashboard</h1>
        <p><strong>Report Period:</strong> {week_ago.strftime("%Y-%m-%d")} to {now.strftime("%Y-%m-%d")}</p>
        
        <div class="dashboard">
            <div class="card highlight">
                <h3>{stats['total']}</h3>
                <p>Total Tenders</p>
            </div>
            <div class="card">
                <h3>{stats['this_week']}</h3>
                <p>Added This Week</p>
            </div>
            <div class="card">
                <h3>{len(stats['closing_soon'])}</h3>
                <p>Closing Soon</p>
            </div>
            <div class="card">
                <h3>{len(stats['high_priority'])}</h3>
                <p>High Priority Open</p>
            </div>
        </div>
        
        <h2>🎯 Priority Distribution</h2>
        <div class="priority-grid">
            <div class="priority-card high">
                <h3>{stats['by_priority']['HIGH']}</h3>
                <p>🔥 HIGH</p>
            </div>
            <div class="priority-card medium">
                <h3>{stats['by_priority']['MEDIUM']}</h3>
                <p>✅ MEDIUM</p>
            </div>
            <div class="priority-card low">
                <h3>{stats['by_priority']['LOW']}</h3>
                <p>📝 LOW</p>
            </div>
        </div>
        
        <h2>📁 By Category</h2>
        <div class="chart-container">
    """
    
    max_type = max(stats['by_type'].values()) if stats['by_type'] else 1
    for t_type, count in stats['by_type'].items():
        width = int((count / max_type) * 100) if max_type > 0 else 0
        html += f'<div class="bar" style="width: {max(width, 10)}%">{t_type}: {count}</div>'
    
    html += """
        </div>
        
        <h2>⚠️ Closing Soon (Next 7 Days)</h2>
    """
    
    if stats['closing_soon']:
        html += """
        <table>
            <tr><th>Ref</th><th>Tender</th><th>Client</th><th>Days Left</th><th>Priority</th></tr>
        """
        for t in stats['closing_soon']:
            urgent = "urgent" if t['days_left'] <= 3 else ""
            badge = "badge-high" if t['priority'] == "HIGH" else "badge-medium"
            html += f"""
            <tr class="{urgent}">
                <td>{t['ref']}</td>
                <td>{t['title']}...</td>
                <td>{t['client']}</td>
                <td><strong>{t['days_left']} days</strong></td>
                <td><span class="badge {badge}">{t['priority']}</span></td>
            </tr>
            """
        html += "</table>"
    else:
        html += "<p>No tenders closing in the next 7 days.</p>"
    
    html += """
        <h2>🔥 Top High Priority Opportunities</h2>
    """
    
    if stats['high_priority']:
        html += """
        <table>
            <tr><th>Ref</th><th>Tender</th><th>Client</th><th>Type</th><th>Score</th></tr>
        """
        for t in stats['high_priority'][:5]:
            type_badge = "badge-tes" if t['type'] == "TES" else "badge-phakathi"
            html += f"""
            <tr>
                <td>{t['ref']}</td>
                <td>{t['title']}...</td>
                <td>{t['client']}</td>
                <td><span class="badge {type_badge}">{t['type']}</span></td>
                <td><strong>{t['score']}/10</strong></td>
            </tr>
            """
        html += "</table>"
    else:
        html += "<p>No high priority open tenders.</p>"
    
    html += """
        <h2>🏭 Top Industries</h2>
        <div class="chart-container">
    """
    
    max_ind = max(stats['top_industries'].values()) if stats['top_industries'] else 1
    for ind, count in stats['top_industries'].items():
        width = int((count / max_ind) * 100) if max_ind > 0 else 0
        html += f'<div class="bar" style="width: {max(width, 10)}%">{ind}: {count}</div>'
    
    html += f"""
        </div>
        
        <hr>
        <p style="color: #666; font-size: 0.8em; text-align: center;">
            Generated by TenderScan AI Engine | {now.strftime("%Y-%m-%d %H:%M")}
        </p>
    </body>
    </html>
    """
    
    return html


def save_weekly_report(html: str) -> str:
    """Save weekly report as HTML file"""
    
    os.makedirs(REPORTS_DIR, exist_ok=True)
    
    filename = f"weekly_report_{datetime.now().strftime('%Y%m%d')}.html"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)
    
    return filepath


def send_weekly_email(html: str, report_path: str) -> bool:
    """Send weekly report via email with attachment"""
    
    if not EMAIL_ENABLED:
        print("📧 Email disabled - skipping")
        return False
    
    if not SMTP_USER or not SMTP_PASSWORD or not EMAIL_TO:
        print("❌ Email configuration incomplete")
        return False
    
    try:
        msg = MIMEMultipart()
        msg["Subject"] = f"📊 Weekly Tender Dashboard - {datetime.now().strftime('%Y-%m-%d')}"
        msg["From"] = EMAIL_FROM
        msg["To"] = ", ".join(EMAIL_TO)
        
        msg.attach(MIMEText(html, "html"))
        
        if os.path.exists(EXCEL_PATH):
            with open(EXCEL_PATH, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", "attachment; filename=Tender_Dashboard.xlsx")
                msg.attach(part)
        
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())
        
        print(f"✅ Weekly report sent to {', '.join(EMAIL_TO)}")
        return True
        
    except Exception as e:
        print(f"❌ Email failed: {e}")
        return False


def run_weekly():
    """Main weekly report function"""
    
    print(f"\n📊 TenderScan Weekly Report - {datetime.now().strftime('%Y-%m-%d')}")
    print("=" * 50)
    
    print("\n📈 Extracting statistics...")
    stats = get_weekly_stats()
    
    if not stats:
        print("❌ No data available")
        return
    
    print("📝 Generating report...")
    html = generate_weekly_html(stats)
    
    report_path = save_weekly_report(html)
    print(f"💾 Report saved: {report_path}")
    
    send_weekly_email(html, report_path)
    
    print("\n" + "=" * 50)
    print("📊 WEEKLY SUMMARY")
    print(f"   Total tenders:    {stats['total']}")
    print(f"   Added this week:  {stats['this_week']}")
    print(f"   Closing soon:     {len(stats['closing_soon'])}")
    print(f"   HIGH priority:    {stats['by_priority']['HIGH']}")
    print(f"   TES tenders:      {stats['by_type']['TES']}")
    print(f"   Phakathi tenders: {stats['by_type']['Phakathi']}")
    print("=" * 50)
    
    return stats


if __name__ == "__main__":
    run_weekly()

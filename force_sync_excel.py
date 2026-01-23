import json
import os
import sys
from datetime import datetime
import openpyxl

# Paths
EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
OUTPUT_JSON = "/Users/lazolasonqishe/Documents/tender-intelligence/output/new_tenders.json"

def export_excel_to_dashboard():
    print(f"Reading from: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    tenders = []
    
    # Headers map (checking inspect output)
    # 0: Tender Name (A) -> title
    # 1: Client (B) -> client
    # 3: Industry (D) -> source
    # 6: Priority (G) -> scores.priority
    # 5: Composite Score (F) -> scores.composite_score
    # 10: Stage (K)
    # 11: Closing Date (L) -> closing_date
    # 15: Ref (P) -> ref
    # 2: Type (C) -> category
    
    for row in range(2, ws.max_row + 1):
        # Read cells
        tender_name = ws.cell(row=row, column=1).value
        client = ws.cell(row=row, column=2).value
        category = ws.cell(row=row, column=3).value
        industry = ws.cell(row=row, column=4).value
        comp_score = ws.cell(row=row, column=6).value
        priority = ws.cell(row=row, column=7).value
        closing_date = ws.cell(row=row, column=12).value
        ref = ws.cell(row=row, column=16).value
        
        if not ref or not tender_name:
            continue
            
        # Parse title from Tender Name "REF - Title"
        title = tender_name
        if " - " in str(tender_name):
            parts = str(tender_name).split(" - ", 1)
            if len(parts[0]) < 50: # likely a ref prefix
                title = parts[1]
        
        # Parse source from Industry "Source (Industry)"
        source = str(industry)
        if "(" in source:
            source = source.split("(")[0].strip()
            
        # Generate URL based on source
        url = ""
        ref_str = str(ref)
        title_str = str(title)[:50]
        if "Eskom" in source:
            url = "https://www.eskom.co.za/eskom-tenders/"
        elif "Rand Water" in source:
            url = "https://www.randwater.co.za/availabletenders.php"
        elif "Johannesburg Water" in source:
            url = "https://www.johannesburgwater.co.za/tenders/"
        elif "National Treasury" in source:
            from urllib.parse import quote
            url = f"https://www.etenders.gov.za/Home/opportunities?TextSearch={quote(ref_str)}"
        elif "Umgeni" in source:
            url = "https://www.umngeni-uthukela.co.za/tender/"
        elif "Magalies" in source:
            url = "https://magalieswater.co.za/tenders/"
        elif "Lepelle" in source:
            url = "https://lepellewater.co.za/procurement/tenders/"
        else:
            from urllib.parse import quote
            url = f"https://www.google.com/search?q={quote(ref_str + ' ' + title_str + ' tender')}"
            
        t_obj = {
            "ref": str(ref),
            "title": str(title),
            "description": str(title), # fallback
            "client": str(client) if client else "Unknown",
            "category": str(category) if category else "General",
            "source": source,
            "closing_date": str(closing_date) if closing_date else "",
            "scores": {
                "priority": str(priority) if priority else "MEDIUM",
                "composite_score": float(comp_score) if comp_score else 0,
                "composite": float(comp_score) if comp_score else 0
            },
            "url": url
        }
        
        tenders.append(t_obj)
        
    print(f"Found {len(tenders)} tenders in Excel.")
    
    # Write to JSON
    data = {
        "meta": {
            "last_sync": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "source": "Excel Export"
        },
        "tenders": tenders
    }
    
    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, "w") as f:
        json.dump(data, f, indent=4)
        
    print(f"Exported to {OUTPUT_JSON}")

if __name__ == "__main__":
    export_excel_to_dashboard()

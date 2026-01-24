#!/usr/bin/env python3
"""
Audit script to show WHY tenders are being excluded.
Shows actual tender titles + exclusion reasons.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from classify_engine import classify_tender
from keyword_rules import NEGATIVE_KEYWORDS, STRONG_MATCH_KEYWORDS

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
SHEET_NAME = "Tender_Log"

def audit_excel_exclusions(limit=50):
    """Audit recent tenders in Excel to see why they were excluded."""

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if not rows:
        print("No rows found in Excel")
        return

    headers = rows[0]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers or []) if h}

    print(f"\n{'='*100}")
    print(f"EXCLUSION AUDIT - Showing why tenders don't appear on dashboard")
    print(f"{'='*100}\n")

    mexel_count = 0
    excluded_count = 0
    samples_shown = 0

    excluded_samples = []
    mexel_samples = []

    for row in rows[1:]:  # Skip header
        if not row or samples_shown >= limit:
            break

        title_idx = header_map.get("Tender Name", 0)
        ref_idx = header_map.get("Reference Number")
        client_idx = header_map.get("Client", 1)

        title = str(row[title_idx] if title_idx < len(row) else "").strip()
        ref = str(row[ref_idx] if ref_idx and ref_idx < len(row) else "").strip()
        client = str(row[client_idx] if client_idx < len(row) else "").strip()

        if not title:
            continue

        # Classify
        result = classify_tender(title, title)
        category = result.get("category", "UNKNOWN")
        reason = result.get("reason", "")

        if category == "MEXEL":
            mexel_count += 1
            mexel_samples.append({
                "ref": ref or "N/A",
                "title": title[:80],
                "client": client,
                "reason": reason
            })
        else:
            excluded_count += 1
            excluded_samples.append({
                "ref": ref or "N/A",
                "title": title[:80],
                "client": client,
                "reason": reason
            })

        samples_shown += 1

    # Show results
    print(f"📊 SUMMARY (First {samples_shown} tenders analyzed):")
    print(f"   ✅ MEXEL-classified: {mexel_count}")
    print(f"   ❌ EXCLUDED: {excluded_count}")
    print(f"   Hit rate: {(mexel_count/samples_shown*100):.1f}%\n")

    if mexel_samples:
        print(f"\n✅ MEXEL TENDERS ({len(mexel_samples)}):")
        print(f"{'-'*100}")
        for i, t in enumerate(mexel_samples[:10], 1):
            print(f"{i}. [{t['ref']}] {t['client']}")
            print(f"   {t['title']}")
            print(f"   ✓ {t['reason']}\n")

    if excluded_samples:
        print(f"\n❌ EXCLUDED TENDERS ({len(excluded_samples)}) - Sample of first 20:")
        print(f"{'-'*100}")
        for i, t in enumerate(excluded_samples[:20], 1):
            print(f"{i}. [{t['ref']}] {t['client']}")
            print(f"   {t['title']}")
            print(f"   ✗ {t['reason']}\n")

    # Categorize exclusion reasons
    print(f"\n📈 EXCLUSION BREAKDOWN:")
    print(f"{'-'*100}")

    reason_counts = {}
    for t in excluded_samples:
        reason = t['reason']
        if "'" in reason:
            # Extract keyword from "Excluded: 'maintenance'"
            keyword = reason.split("'")[1] if "'" in reason else reason
        else:
            keyword = reason
        reason_counts[keyword] = reason_counts.get(keyword, 0) + 1

    for keyword, count in sorted(reason_counts.items(), key=lambda x: x[1], reverse=True)[:15]:
        print(f"   {keyword}: {count} tenders")

    print(f"\n{'='*100}")
    print(f"RECOMMENDATION:")
    print(f"{'='*100}")

    if excluded_count > mexel_count * 5:
        print("⚠️  FILTERS ARE TOO STRICT")
        print("   Over 80% of tenders are being excluded.")
        print("   Common exclusion keywords like 'maintenance', 'repair' should be reviewed.")
        print("\nSuggested fixes:")
        print("   1. Remove 'maintenance' from NEGATIVE_KEYWORDS")
        print("   2. Remove 'repair' from NEGATIVE_KEYWORDS")
        print("   3. Add 'preventative maintenance' as specific exclusion instead")
        print("   4. Expand CONTEXT_KEYWORDS to catch more water treatment tenders")
    else:
        print("✅ Filter balance looks reasonable")

if __name__ == "__main__":
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 100
    audit_excel_exclusions(limit)

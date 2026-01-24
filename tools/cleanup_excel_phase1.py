#!/usr/bin/env python3
"""
PHASE 1: Excel Deep Clean
Removes corrupted rows (8-147) with classification output in Reference Number field.
Preserves valid historical tenders (rows 2-7) and active tenders (rows 148-150).
Updates schema to 15 columns (removes Revenue Potential, Risk Level).
"""

import openpyxl
import shutil
import os
import sys
from datetime import datetime

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
BACKUP_PATH = f"/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2_PHASE1_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
SHEET_NAME = "Tender_Log"

NEW_HEADERS = [
    "Tender Name", "Client", "Type", "Industry", "Fit Score",
    "Composite Score", "Priority", "Mexel Fit", "Stage",
    "Closing Date", "Status", "Next Action", "Notes",
    "Reference Number", "Date Added"
]

def is_corrupted_row(row_values):
    """
    Detect corrupted rows by checking Reference Number column (index 15 in old schema).
    Corrupted rows contain classification output like:
    - "TES override keyword detected"
    - "[AI Score: 8.8/10"
    - "Phakathi override"
    """
    if len(row_values) < 16:
        return False

    ref = row_values[15]  # Old schema: Reference Number was column 16 (index 15)
    if not ref:
        return False

    ref_str = str(ref).lower()
    corruption_markers = [
        "tes override",
        "phakathi",
        "[ai score",
        "priority bid",
        "no clear classification"
    ]

    return any(marker in ref_str for marker in corruption_markers)

def cleanup(live=False):
    """Main cleanup function"""

    if not live:
        print("="*80)
        print("DRY RUN MODE - No changes will be saved")
        print("="*80)
    else:
        print("="*80)
        print("LIVE MODE - Creating backup and cleaning Excel")
        print("="*80)
        print(f"Backup: {BACKUP_PATH}")
        shutil.copy2(EXCEL_PATH, BACKUP_PATH)
        print("✅ Backup created")

    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    print(f"\nCurrent Excel state:")
    print(f"  Total rows: {ws.max_row}")
    print(f"  Total columns: {ws.max_column}")

    # Scan for corrupted rows
    corrupted_rows = []
    valid_rows = []

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, 20)]

        if is_corrupted_row(row_values):
            corrupted_rows.append(row_idx)
        else:
            # Check if row has any data
            if any(v for v in row_values if v):
                valid_rows.append(row_idx)

    print(f"\n📊 Analysis:")
    print(f"  Corrupted rows: {len(corrupted_rows)}")
    print(f"  Valid rows: {len(valid_rows)}")

    if corrupted_rows:
        print(f"\n  Corrupted row range: {corrupted_rows[0]}-{corrupted_rows[-1]}")

    if len(valid_rows) > 0:
        print(f"\n✅ Valid rows to preserve:")
        for row_idx in valid_rows[:10]:  # Show first 10
            ref = ws.cell(row=row_idx, column=16).value
            title = ws.cell(row=row_idx, column=1).value
            closing = ws.cell(row=row_idx, column=12).value
            print(f"    Row {row_idx}: {ref} - {str(title)[:50]} (closes: {closing})")

        if len(valid_rows) > 10:
            print(f"    ... and {len(valid_rows) - 10} more")

    # Delete corrupted rows (in reverse order to maintain indices)
    if corrupted_rows and live:
        print(f"\n🗑️  Deleting {len(corrupted_rows)} corrupted rows...")
        for row_idx in reversed(corrupted_rows):
            ws.delete_rows(row_idx, 1)
        print("✅ Corrupted rows deleted")

    # Update headers to 15-column schema
    print(f"\n📝 Updating headers to 15-column schema...")
    for i, header in enumerate(NEW_HEADERS, 1):
        ws.cell(row=1, column=i).value = header

    # Clear old column headers (16-17 from old 17-column schema)
    ws.cell(row=1, column=16).value = None
    ws.cell(row=1, column=17).value = None

    # Update all data rows to 15-column schema (remove cols 9-10: Risk Level, Revenue Potential)
    print(f"📝 Remapping columns for {len(valid_rows)} rows...")

    for row_idx in range(2, ws.max_row + 1):
        # Read old schema (17 columns)
        old_row = [ws.cell(row=row_idx, column=c).value for c in range(1, 18)]

        # Skip empty rows
        if not any(old_row):
            continue

        # Map to new schema (15 columns)
        # Old: Name, Client, Type, Industry, Fit, Composite, Priority, Mexel, Risk, Revenue, Stage, Closing, Status, Action, Notes, Ref, Date
        # New: Name, Client, Type, Industry, Fit, Composite, Priority, Mexel, Stage, Closing, Status, Action, Notes, Ref, Date

        new_row = [
            old_row[0],   # Tender Name
            old_row[1],   # Client
            old_row[2],   # Type
            old_row[3],   # Industry
            old_row[4],   # Fit Score
            old_row[5],   # Composite Score
            old_row[6],   # Priority
            old_row[7],   # Mexel Fit
            old_row[10],  # Stage (skip Risk Level [8], Revenue Potential [9])
            old_row[11],  # Closing Date
            old_row[12],  # Status
            old_row[13],  # Next Action
            old_row[14],  # Notes
            old_row[15],  # Reference Number
            old_row[16],  # Date Added
        ]

        # Write new schema
        for col_idx, value in enumerate(new_row, 1):
            ws.cell(row=row_idx, column=col_idx).value = value

        # Clear old columns
        ws.cell(row=row_idx, column=16).value = None
        ws.cell(row=row_idx, column=17).value = None

    print("✅ Column remapping complete")

    # Final state
    print(f"\n📊 Final Excel state:")
    print(f"  Total rows: {ws.max_row} (1 header + {ws.max_row - 1} data)")
    print(f"  Total columns: 15")

    # Save if live
    if live:
        wb.save(EXCEL_PATH)
        print(f"\n✅ Changes saved to: {EXCEL_PATH}")
        print(f"✅ Backup available at: {BACKUP_PATH}")
    else:
        print(f"\n⚠️  DRY RUN complete - No changes saved")
        print(f"\nTo apply changes, run: python3 {__file__} --live")

    return True

if __name__ == "__main__":
    is_live = "--live" in sys.argv
    cleanup(live=is_live)

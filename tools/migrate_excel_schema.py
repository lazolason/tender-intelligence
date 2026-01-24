#!/usr/bin/env python3
import openpyxl
import shutil
import os
import sys
from datetime import datetime

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
BACKUP_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2_MIGRATION_BACKUP.xlsx"
SHEET_NAME = "Tender_Log"

NEW_HEADERS = [
    "Tender Name", "Client", "Type", "Industry", "Fit Score", 
    "Composite Score", "Priority", "Mexel Fit", "Stage",
    "Closing Date", "Status", "Next Action", "Notes",
    "Reference Number", "Date Added"
]

def migrate(live=False):
    if not live:
        print("--- DRY RUN MODE (No changes will be saved) ---")
    else:
        print("--- LIVE MIGRATION START ---")
        print(f"Creating backup at: {BACKUP_PATH}")
        shutil.copy2(EXCEL_PATH, BACKUP_PATH)
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    
    # 1. Update Headers
    print(f"Updating headers...")
    for i, h in enumerate(NEW_HEADERS, 1):
        ws.cell(row=1, column=i).value = h

    # 2. Process Rows
    old_rows_count = 0
    new_rows_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        raw_row = [ws.cell(row=row_idx, column=c).value for c in range(1, 19)]
        filled_cols = [i for i, v in enumerate(raw_row) if v is not None]
        last_filled_idx = max(filled_cols) if filled_cols else 0
        
        if last_filled_idx <= 10:
            old_rows_count += 1
            old = raw_row[:11] 
            
            # Re-map to new positions
            # Clear row
            for c in range(1, 19): ws.cell(row=row_idx, column=c).value = None
            
            ws.cell(row=row_idx, column=1).value = old[0]  # Name
            ws.cell(row=row_idx, column=2).value = old[1]  # Client
            ws.cell(row=row_idx, column=3).value = old[2]  # Type
            ws.cell(row=row_idx, column=4).value = old[3]  # Industry
            ws.cell(row=row_idx, column=5).value = old[4]  # Fit Score
            ws.cell(row=row_idx, column=6).value = old[4]  # Composite (Default)
            ws.cell(row=row_idx, column=7).value = "MEDIUM" # Priority (Default)
            ws.cell(row=row_idx, column=11).value = old[5] # Stage
            ws.cell(row=row_idx, column=12).value = old[6] # Closing Date
            ws.cell(row=row_idx, column=13).value = old[7] # Status
            ws.cell(row=row_idx, column=14).value = old[8] # Next Action
            ws.cell(row=row_idx, column=15).value = old[9] # Notes
            
            if old[0] and " - " in str(old[0]):
                ws.cell(row=row_idx, column=16).value = str(old[0]).split(" - ")[0]
            
            ws.cell(row=row_idx, column=17).value = old[10] # Date Added
        else:
            new_rows_count += 1

    print(f"Summary: {old_rows_count} rows migrated, {new_rows_count} rows preserved.")
    
    if live:
        wb.save(EXCEL_PATH)
        print("✅ Live changes saved to Excel.")
    else:
        print("⚠️ Dry run complete. No changes saved.")

if __name__ == "__main__":
    is_live = "--live" in sys.argv
    migrate(live=is_live)

import openpyxl
import os
import sys

# Path from config
EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"

def inspect():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ File not found at: {EXCEL_PATH}")
        return

    print(f"✅ Loading: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        print(f"Rows: {ws.max_row}")
        
        print("\nLast 10 entries:")
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        for row in range(max(2, ws.max_row - 9), ws.max_row + 1):
            vals = [ws.cell(row=row, column=c).value for c in range(1, 18)]
            # 1=Name, 12=Closing, 16=Ref
            ref = vals[15]
            name = vals[0]
            date = vals[11]
            print(f"Row {row}: {ref} | {name} | {date}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect()

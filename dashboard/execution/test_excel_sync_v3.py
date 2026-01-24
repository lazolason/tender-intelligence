import openpyxl
from datetime import datetime
import os

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
SHEET_NAME = "Tender_Log"

def test_read_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb[SHEET_NAME]
    headers = [cell.value for cell in sheet[1] if cell.value]
    
    max_row = sheet.max_row
    start_row = max(2, max_row - 20)
    
    print(f"Reading from row {start_row} to {max_row}")
    today = datetime.now().date()
    found = 0
    
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
        if not any(row): continue
        data = dict(zip(headers, row))
        
        closing = data.get('Closing Date')
        # Handle string dates like '2026-02-15' or '15/02/2026'
        closing_dt = None
        if isinstance(closing, datetime):
            closing_dt = closing.date()
        elif isinstance(closing, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d %b %Y"]:
                try:
                    closing_dt = datetime.strptime(closing, fmt).date()
                    break
                except: continue
        
        if closing_dt and closing_dt >= today:
            found += 1
            print(f"ACTIVE: {data.get('Tender Name')} | Closing: {closing_dt}")
            
    print(f"Total active found in last 20 rows: {found}")

if __name__ == "__main__":
    test_read_excel()

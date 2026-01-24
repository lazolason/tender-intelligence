import openpyxl
import os

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
SHEET_NAME = "Tender_Log"

def test_read_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb[SHEET_NAME]
    headers = [cell.value for cell in sheet[1] if cell.value]
    
    max_row = sheet.max_row
    start_row = max(2, max_row - 10)
    
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
        if not any(row): continue
        data = dict(zip(headers, row))
        print(f"Ref/Title: {data.get('Tender Name')} | Raw Closing Date: {data.get('Closing Date')} | Type: {type(data.get('Closing Date'))}")

if __name__ == "__main__":
    test_read_excel()

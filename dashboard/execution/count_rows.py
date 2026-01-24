import openpyxl
import os

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Tender_Log"]

old_count = 0
new_count = 0

for row_idx in range(2, ws.max_row + 1):
    row = [ws.cell(row=row_idx, column=c).value for c in range(1, 19)]
    filled_cols = [i for i, v in enumerate(row) if v is not None]
    last_filled_idx = max(filled_cols) if filled_cols else 0
    
    if last_filled_idx <= 10:
        old_count += 1
    else:
        new_count += 1

print(f"VERIFICATION: {old_count} old rows, {new_count} new rows.")

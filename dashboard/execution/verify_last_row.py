import openpyxl
import os

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
SHEET_NAME = "Tender_Log"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb[SHEET_NAME]
max_row = ws.max_row
print(f"Total Rows: {max_row}")

# Check the very last row
last_row_vals = [c.value for c in ws[max_row]]
print(f"Last Row ({max_row}): {last_row_vals}")

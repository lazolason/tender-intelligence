import openpyxl
import os

EXCEL_PATH = "/Users/lazolasonqishe/Documents/MASTER/TENDERS/01_Tender_Log/Tender_Dashboard_v2.xlsx"
SHEET_NAME = "Tender_Log"

if os.path.exists(EXCEL_PATH):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        headers = [c.value for c in ws[1]]
        print(f"Headers: {headers}")
        
        # Check first data row
        if ws.max_row > 1:
            row_2 = [c.value for c in ws[2]]
            print(f"Row 2: {row_2}")
    else:
        print(f"Sheet {SHEET_NAME} missing")
else:
    print("Excel file missing")

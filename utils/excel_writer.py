# ==========================================================
# EXCEL WRITER UTILITY
# Writes tender data to Excel with scoring columns
# ==========================================================

import os
import sys
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging
import re

# Assuming these are in the parent directory, adjust if necessary
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from classify_engine import classify_tender
from scoring_engine import score_tender
from utils.duplicate_detector import find_duplicate, find_best_title_match

logger = logging.getLogger(__name__)
# Column headers (with new scoring columns)
HEADERS = [
    "Tender Name",
    "Client",
    "Type",
    "Industry",
    "Fit Score",
    "Composite Score",
    "Priority",
    "TES Fit",
    "Risk Level",
    "Revenue Potential",
    "Stage",
    "Closing Date",
    "Status",
    "Next Action",
    "Notes",
    "Reference Number",
    "Date Added"
]

# Priority colors
PRIORITY_COLORS = {
    "HIGH": "FF6B6B",     # Red
    "MEDIUM": "FFE66D",   # Yellow
    "LOW": "C8E6C9"       # Green
}


class ExcelWriter:
    """
    Writes tender data to Excel spreadsheet with scoring.
    
    Manages Excel workbook operations including:
    - Loading existing workbooks or creating new ones
    - Writing tender data with duplicate detection
    - Applying scoring and formatting
    - Maintaining caches for performance
    
    Attributes:
        file_path: Path to Excel file
        sheet_name: Name of worksheet
        log_file_path: Path to log file for logging
        fuzzy_duplicate_threshold: Threshold for fuzzy duplicate detection (0-100)
        fuzzy_date_window_days: Days window for date-based duplicate matching
    """
    
    def __init__(
        self,
        file_path: str,
        sheet_name: str = "Tender_Log",
        *,
        log_file_path: str = None,
        fuzzy_duplicate_threshold: int = 85,
        fuzzy_date_window_days: int = 7,
    ) -> None:
        """
        Initialize Excel writer.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of worksheet (default: "Tender_Log")
            log_file_path: Optional path to log file
            fuzzy_duplicate_threshold: Fuzzy match threshold (default: 85)
            fuzzy_date_window_days: Date window for duplicates (default: 7)
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.log_file_path = log_file_path
        self.fuzzy_duplicate_threshold = int(fuzzy_duplicate_threshold)
        self.fuzzy_date_window_days = int(fuzzy_date_window_days)
        self._existing_refs_cache = None
        self._existing_tenders_cache = None
        self._ensure_workbook()

    def _log(self, message: str, level: str = "INFO") -> None:
        """
        Log message to file and console

        Args:
            message: Message to log
            level: Log level (default: "INFO")
        """
        if self.log_file_path:
            try:
                from utils.logging_tools import write_log

                write_log(self.log_file_path, message, level)
                return
            except Exception:
                pass

        print(message)

    def _ensure_workbook(self) -> None:
        """
        Create workbook if it doesn't exist, otherwise load existing one
        
        Side effects:
            Sets self.wb to the workbook object
            Creates headers if new workbook
            Saves the workbook to file
        """
        if os.path.exists(self.file_path):
            self.wb = load_workbook(self.file_path)
        else:
            self.wb = Workbook()
            self.wb.active.title = self.sheet_name
            self._write_headers()
            self.wb.save(self.file_path)
    
    def _write_headers(self) -> None:
        """
        Write column headers with formatting
        
        Side effects:
            Writes headers to first row of worksheet
            Applies font, fill, alignment, and column width formatting
        """
        ws = self.wb.active
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        col_widths = {
            'A': 40,  # Tender Name
            'B': 20,  # Client
            'C': 12,  # Type
            'D': 25,  # Industry
            'E': 10,  # Fit Score
            'F': 14,  # Composite Score
            'G': 10,  # Priority
            'H': 10,  # TES Fit
            'I': 12,  # Risk Level
            'J': 15,  # Revenue Potential
            'K': 10,  # Stage
            'L': 12,  # Closing Date
            'M': 10,  # Status
            'N': 15,  # Next Action
            'O': 50,  # Notes
            'P': 20,  # Reference Number
            'Q': 12,  # Date Added
        }
        
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    def _get_existing_references(self):
        """Get set of existing tender reference numbers"""
        if self._existing_refs_cache is not None:
            return self._existing_refs_cache

        ws = self.wb.active
        existing = set()
        
        for row in range(2, ws.max_row + 1):
            ref = ws.cell(row=row, column=16).value  # Reference Number column
            if ref:
                existing.add(str(ref).strip().upper())
        
        self._existing_refs_cache = existing
        return existing

    def _extract_title_from_tender_name(self, tender_name: str) -> str:
        if not tender_name:
            return ""
        # Common format: "REF - Title"
        parts = str(tender_name).split(" - ", 1)
        if len(parts) == 2 and len(parts[0]) <= 60:
            return parts[1].strip()
        return str(tender_name).strip()

    def _extract_source_from_industry_cell(self, industry_value: str) -> str:
        if not industry_value:
            return "Unknown"
        value = str(industry_value).strip()
        # Stored as: "{source} ({industry_matched})"
        match = re.match(r"^(.*?)\s*\(", value)
        return (match.group(1) if match else value).strip() or "Unknown"

    def _get_existing_tenders_metadata(self) -> List[Dict[str, any]]:
        """
        Get metadata for existing tenders from Excel sheet
        
        Returns:
            List of tender metadata dictionaries with ref, title, source, closing_date
        """
        if self._existing_tenders_cache is not None:
            return self._existing_tenders_cache
 
        ws = self.wb.active
        tenders = []
 
        for row in range(2, ws.max_row + 1):
            tender_name = ws.cell(row=row, column=1).value
            industry = ws.cell(row=row, column=4).value
            closing_date = ws.cell(row=row, column=12).value
            ref = ws.cell(row=row, column=16).value
 
            ref_norm = str(ref).strip().upper() if ref else ""
            title = self._extract_title_from_tender_name(tender_name)
            source = self._extract_source_from_industry_cell(industry)
 
            if not title and not ref_norm:
                continue
 
            tenders.append(
                {
                    "ref": ref_norm,
                    "title": title,
                    "source": source,
                    "closing_date": str(closing_date).strip() if closing_date else "",
                }
            )
 
        self._existing_tenders_cache = tenders
        return tenders
    
    def write_tender(self, tender_name: str, client: str, tender_type: str,
                    industry: str, fit_score: int, stage: str, closing_date: str,
                    status: str, next_action: str, notes: str, reference_number: str,
                    composite_score: float = None, priority: str = None,
                    risk_level: str = None, revenue_potential: str = None,
                    tes_fit: int = None) -> bool:
        """
        Write a single tender to Excel
        
        Args:
            tender_name: Full tender name (e.g., "REF-001 - Title")
            client: Client organization name
            tender_type: Tender category/classification
            industry: Industry type with source
            fit_score: TES fit score (0-10)
            stage: Tender stage (e.g., "New", "In Progress")
            closing_date: Tender closing date string
            status: Current status (e.g., "Open", "Closed")
            next_action: Recommended next action
            notes: Additional notes
            reference_number: Tender reference number
            composite_score: Overall composite score (0-10)
            priority: Priority level (HIGH, MEDIUM, LOW)
            risk_level: Risk assessment (Low, Medium, High)
            revenue_potential: Revenue potential (Low, Medium, High)
            tes_fit: TES suitability score (0-10)
            
        Returns:
            True if tender was added, False if duplicate
        """
        
        # Check for duplicates
        existing = self._get_existing_references()
        ref_normalized = str(reference_number).strip().upper()
        
        if ref_normalized and ref_normalized != "NA" and ref_normalized in existing:
            return False  # Duplicate
        
        # Also check by tender name
        ws = self.wb.active
        for row in range(2, ws.max_row + 1):
            existing_name = ws.cell(row=row, column=1).value
            if existing_name and existing_name.strip().upper() == tender_name.strip().upper():
                return False  # Duplicate
        
        # Add new row
        row = ws.max_row + 1
        
        data = [
            tender_name,
            client,
            tender_type,
            industry,
            fit_score,
            composite_score or fit_score,
            priority or "MEDIUM",
            tes_fit or 0,
            risk_level or "Medium",
            revenue_potential or "Medium",
            stage,
            closing_date,
            status,
            next_action,
            notes,
            reference_number,
            datetime.now().strftime("%Y-%m-%d")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            
            # Apply priority color to the row
            if priority and priority in PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=PRIORITY_COLORS[priority],
                    end_color=PRIORITY_COLORS[priority],
                    fill_type="solid"
                )
        
        # Save workbook
        self.wb.save(self.file_path)

        # Update caches (fast path)
        if self._existing_refs_cache is not None and ref_normalized:
            self._existing_refs_cache.add(ref_normalized)
        if self._existing_tenders_cache is not None:
            self._existing_tenders_cache.append(
                {
                    "ref": ref_normalized,
                    "title": self._extract_title_from_tender_name(tender_name),
                    "source": self._extract_source_from_industry_cell(industry),
                    "closing_date": str(closing_date).strip() if closing_date else "",
                }
            )
        return True

    def add_tender_with_scoring(self, tender_data: dict) -> Tuple[bool, Dict[str, Any], Dict[str, Any]]:
        """
        Score a tender and write it to Excel file.
        
        Args:
            tender_data: Dictionary containing tender information with keys:
                - ref: Reference number
                - title: Tender title
                - description: Tender description
                - client: Client organization
                - closing_date: Closing date string
                - source: Source name
                
        Returns:
            Tuple of (was_added, scores_dict, classification_dict)
            - was_added: True if tender was written, False if duplicate
            - scores_dict: Dictionary with scoring results
            - classification_dict: Dictionary with classification results
            
        Side effects:
            - Writes to Excel file if not duplicate
            - Updates internal caches
            - Logs duplicate warnings
        """
        # Fuzzy duplicate check (skip scoring if likely duplicate)
        try:
            existing_refs = self._get_existing_references()
            ref_normalized = str(tender_data.get("ref", "")).strip().upper()
            if ref_normalized and ref_normalized != "NA" and ref_normalized in existing_refs:
                self._log(f"[DUPLICATE] Exact ref match: {ref_normalized}", "INFO")
                return False, {}, {"category": "Unknown", "reason": "", "short_title": "Tender"}

            existing_meta = self._get_existing_tenders_metadata()
            match = find_duplicate(
                {
                    "ref": tender_data.get("ref", ""),
                    "title": tender_data.get("title", ""),
                    "source": tender_data.get("source", "Unknown"),
                    "closing_date": tender_data.get("closing_date", ""),
                },
                existing_meta,
                threshold=self.fuzzy_duplicate_threshold,
                date_window_days=self.fuzzy_date_window_days,
                require_same_source=True,
            )

            if match and match.is_duplicate:
                new_ref = tender_data.get("ref") or "NA"
                self._log(
                    f"[DUPLICATE] {new_ref}: {match.reason} ({match.similarity}%) vs {match.existing_ref} [{match.existing_source}]",
                    "WARNING",
                )
                return False, {}, {"category": "Unknown", "reason": "", "short_title": "Tender"}

            # Log near-duplicates for review (don’t block add unless it meets duplicate criteria)
            best = find_best_title_match(
                {"title": tender_data.get("title", "")},
                existing_meta,
            )
            if best:
                best_similarity, best_existing = best
                if best_similarity >= self.fuzzy_duplicate_threshold:
                    new_source = (tender_data.get("source") or "Unknown").strip()
                    ex_source = (best_existing.get("source") or "Unknown").strip()
                    if new_source == ex_source:
                        self._log(
                            f"[NEAR-DUPLICATE] {tender_data.get('ref','NA')}: title similarity {best_similarity}% with existing {best_existing.get('ref','NA')} (same source)",
                            "WARNING",
                        )
        except Exception as exc:
            logger.warning("Fuzzy duplicate check failed; continuing: %s", exc)

        # Classify tender (returns dict)
        classification = classify_tender(tender_data["title"], tender_data["description"])
        category = classification["category"]
        reason = classification["reason"]
        short_title = classification["short_title"]

        tender_name = f"{tender_data['ref']} - {tender_data['title']}" if tender_data['ref'] and tender_data['ref'] != "NA" else tender_data['title']

        # AI SCORING ENGINE
        scores = score_tender(
            title=tender_data["title"],
            description=tender_data["description"],
            client=tender_data["client"],
            closing_date=tender_data["closing_date"],
            category=category
        )

        fit_score = scores["fit_score"]
        composite_score = scores["composite_score"]
        priority = scores["priority"]
        recommendation = scores["recommendation"]

        # Build notes with scoring info
        enhanced_notes = f"{reason}\n" if reason else ""
        enhanced_notes += f"[Score: {composite_score}/10 | Priority: {priority}]"
        enhanced_notes += f"\n{recommendation}"

        # Write to Excel
        was_added = self.write_tender(
            tender_name=tender_name,
            client=tender_data["client"],
            tender_type=category,
            industry=f"{tender_data['source']} ({scores['industry_matched']})",
            fit_score=fit_score,
            stage="New",
            closing_date=tender_data["closing_date"],
            status="Open",
            next_action="Review" if priority == "LOW" else "Prepare Bid" if priority == "MEDIUM" else "URGENT BID",
            notes=enhanced_notes,
            reference_number=tender_data["ref"],
            composite_score=composite_score,
            priority=priority,
            risk_level=scores["risk_level"],
            revenue_potential=scores["revenue_potential"],
            tes_fit=scores["tes_suitability"]
        )

        return was_added, scores, classification
    
    def get_stats(self):
        """Get tender statistics"""
        ws = self.wb.active
        
        stats = {
            "total": ws.max_row - 1,  # Exclude header
            "by_type": {},
            "by_priority": {"HIGH": 0, "MEDIUM": 0, "LOW": 0},
            "by_status": {}
        }
        
        for row in range(2, ws.max_row + 1):
            # By type
            t_type = ws.cell(row=row, column=3).value or "Unknown"
            stats["by_type"][t_type] = stats["by_type"].get(t_type, 0) + 1
            
            # By priority
            priority = ws.cell(row=row, column=7).value or "MEDIUM"
            if priority in stats["by_priority"]:
                stats["by_priority"][priority] += 1
            
            # By status
            status = ws.cell(row=row, column=13).value or "Unknown"
            stats["by_status"][status] = stats["by_status"].get(status, 0) + 1
        
        return stats


# ==========================================================
# STANDALONE TEST
# ==========================================================
if __name__ == "__main__":
    # Test
    writer = ExcelWriter("/tmp/test_tender_log.xlsx", "Tender_Log")
    
    added = writer.write_tender(
        tender_name="TEST-001 - Cooling Water Treatment",
        client="Eskom",
        tender_type="MEXEL",
        industry="Power Generation",
        fit_score=8,
        stage="New",
        closing_date="2025-12-15",
        status="Open",
        next_action="Prepare Bid",
        notes="Test tender",
        reference_number="TEST-001",
        composite_score=8.5,
        priority="HIGH",
        risk_level="Low",
        revenue_potential="High",
        tes_fit=9
    )
    
    print(f"Added: {added}")
    print(f"Stats: {writer.get_stats()}")
